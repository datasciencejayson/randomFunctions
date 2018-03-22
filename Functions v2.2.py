# -*- coding: utf-8 -*-
"""
Created on Wed Aug 16 14:31:09 2017

@author: backesj
"""
DIR = 'H:/'
FileType = '.xlsx'

def findFile(DIR, fileType='all', cond='modified', direction='max'): 
    """
    This function will return the file name of the last file created or
    modifed in a given directory. 
    Required arguments:
        DIR         (the directory you want to search)
       
    Optional aruguments:
        fileType    -the type of file you want to find. options are any file
                     extention.
                        (Default is 'all' files)
        cond        -the condition you want to use. options are 'modified',
                     'created', and 'size'
                        (Default is 'modified')
                      be 'created'. 
    """
    # import os package
    import os
    from os import listdir
    from os.path import isfile, join
    chars = set("""~`!@#$%^&*()[]{}_+-=/\|":;?<>,""")
    #import platform
    fileType = ',eed'
    if fileType == 'all':
        file_list = [f for f in listdir(DIR) if isfile(join(DIR, f))]
    else:
         for i, c in enumerate(fileType):
            if c in chars:
                raise TypeError('Check your file type extention')
                
                
                
            elif i == len(fileType):
                fileType = '.'+fileType       
        
        
        if fileType[0] =='.':
            fileType = fileType[fileType.find('.'):]
        for i, c in enumerate(fileType):
            if c in chars:
                raise TypeError('Check your file type extention')
            elif i == len(fileType):
                fileType = '.'+fileType
        file_list = [f for f in listdir(DIR) if isfile(join(DIR, f)) and fileType in f]
        # create file list of xls files in download file directory
        file_list2 = []
        for file in file_list:
            if cond == 'modified':
                # get date of file modificaion in download file directory
                file_list2.append(os.path.getmtime('%s/%s' % (DIR, file)))
            elif cond == 'created':
                # get date of file created in download file directory
                file_list2.append(os.path.getctime('%s/%s' % (DIR, file)))
            elif cond == 'size':
                # get date of file created in download file directory
                file_list2.append(os.path.getsize('%s/%s' % (DIR, file)))            
            else:
                print("Value Error: time argument value must be 'modified' or 'created'")
        # create empty dictionary
        d = {}
        # create dictionary of files and date 
        for i in range(len(file_list)):
            d[file_list[i]] = file_list2[i]
        
        # get max item in dictionary (last item created) this should be the file you
        # downloaded from the website
        import operator
        if direction == 'max':
            return max(d.items(), key=operator.itemgetter(1))[0]
        else:
            return min(d.items(), key=operator.itemgetter(1))[0]


max_value = findFile('H:/','all', 'size', 'max')


def getLatestFile(DIR, FileType):

    #import os packages
    import os
    
    from os import listdir
    from os.path import isfile, join
    #import platform
    file_list = [f for f in listdir(DIR) if isfile(join(DIR, f)) and FileType in f]
    
    # create file list of xls files in download file directory
    file_list2 = []
    for file in file_list:
        # get date of file creation in download file directory
        file_list2.append(os.path.getmtime('%s/%s' % (DIR, file)))
    # create empty dictionary
    d = {}
    # create dictionary of files and date 
    for i in range(len(file_list)):
        d[file_list[i]] = file_list2[i]

    # get max item in dictionary (last item created) this should be the file you
    # downloaded from the website
    import operator
    return max(d.items(), key=operator.itemgetter(1))[0]


max_value = getLatestFile('H:/', '.docx')



def charReplace(inputVar, 
                chars="""~`!@#$%^&*()[]{}_+-=/\|"':;?<>.,""", 
                repace_char=' '):
    
   if type(inputVar) == str:
        for c in chars:
            inputVar = inputVar.replace(c, " ")
        return inputVar
   if type(inputVar) == list:
        print('list')
        outputVar = []
        for string in inputVar:
            print(string)
            for c in chars:
                string = string.replace(c, " ")
            outputVar.append(string)
        return outputVar

charReplace(j, '$' , '_')     
            

     
def findLastFile(DIR, fileType, time='modified'):  
    # import os package
    import os
    
    from os import listdir
    from os.path import isfile, join
    #import platform
    file_list = [f for f in listdir(DIR) if isfile(join(DIR, f)) and fileType in f]
    
    # create file list of xls files in download file directory
    file_list2 = []
    for file in file_list:
        if time == 'modified':
            # get date of file modificaion in download file directory
            file_list2.append(os.path.getmtime('%s/%s' % (DIR, file)))
        elif time == 'created':
z            # get date of file created in download file directory
            file_list2.append(os.path.getctime('%s/%s' % (DIR, file)))
        else:
            print("Value Error: time argument value must be 'modified' or 'created'")
    # create empty dictionary
    d = {}
    # create dictionary of files and date 
    for i in range(len(file_list)):
        d[file_list[i]] = file_list2[i]
    
    # get max item in dictionary (last item created) this should be the file you
    # downloaded from the website
    import operator
    return max(d.items(), key=operator.itemgetter(1))[0]


max_value = getLatestFile(outDIR, '.xlsx')
j = ['asd', 'asd','asd','asd' ]
                
def dedup(seq):
    """
    removes duplicate values from a list or 
    duplicate characters from a string 
    """
    if type(seq) == list:
        seen = set()
        seen_add = seen.add
        return [x for x in seq if not (x in seen or seen_add(x))]    
    elif type(seq) == str:
        seen = set()
        seen_add = seen.add
        return ''.join([x for x in seq if not (x in seen or seen_add(x))])    
    else:
        print("Currently function can only handle lists and strings")
        
dedup(j)
        
        
def column2Number(col):
    import string
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num
    
column2Number('zz')
    
def Number2Column(num):
    import string
    title = ''
    alist = string.ascii_uppercase
    while num:
        mod = (num-1) % 26
        num = int((num - mod) / 26)  
        title += alist[mod]
    return title[::-1]

Number2Column(138)
string.ascii_letters
'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'

def xls2Xlsx(DIR, inFile, outFile):
    
    import pandas as pd
    # DIR name check
    if len(DIR) < 2:
        DIR += ':'
    
    # file type check        
   
    if inFile[-5:] == '.xlsx':
        raise ValueError('File is already in .xlsx format')
    if inFile[-4:] == '.xls':
        try:
            df = pd.read_excel('%s/%s' % (DIR, inFile))
        except FileNotFoundError as detail:
            print("Most likely you have a problem with your inFile variable. \
                  Check to see if file exists. If so, check to see if it is  \
                  actually an '.xls' file not and '.xlsx'. Check the Python \
                  error below for details. \
                  __________________________________________________________\
                  Python error:                                                               "\
                  ,  detail)
        else:
            df = pd.read_excel('%s/%s' % (DIR, inFile))
            writer = pd.ExcelWriter('%s/%s.xlsx' % (DIR, outFile), engine='xlsxwriter')
        
            df.to_excel(writer, index = False)
    
            writer.close() 
    elif inFile[-4:] not in ['xlsx', '.xls']:
        inFile += '.xls'
        try:
            df = pd.read_excel('%s/%s' % (DIR, inFile))
        except FileNotFoundError as detail:
            print("Most likely you have a problem with your inFile variable. \
                  Check to see if file exists. If so, check to see if it is  \
                  actually an '.xls' file not and '.xlsx'. Check the Python \
                  error below for details. \
                  __________________________________________________________\
                  Python error:                                                               "\
                  ,  detail)
        else:
            df = pd.read_excel('%s/%s' % (DIR, inFile))
            writer = pd.ExcelWriter('%s/%s.xlsx' % (DIR, outFile), engine='xlsxwriter')
        
            df.to_excel(writer, index = False)
    
            writer.close() 

xls2Xlsx('h:', 'WV actions.xls', 'New name2')     
    

type('WV actions.xls')
inFile = 'ashfd.xls'
    
    
    
def rename(DIR, inFile, outFile):
    
    import pandas as pd
    
    df = pd.read_excel('%s/%s' % (DIR, inFile))
    
    writer = pd.ExcelWriter('%s/%s.xlsx' % (DIR, outFile), engine='xlsxwriter')
    
    df.to_excel(writer, index = False)

    writer.close()    
DIR = 'h'
len(DIR)
DIR +=':' 
xls2Xlsx('h', a, 'New name')     


def openpyxlFormat(DIR, inFile, outFile):
    wb = load_workbook(filename = '%s/%s.xlsx' % (DIR, inFile))
    sheet_names = wb.get_sheet_names()
    
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

    
    
    def as_text(value):
        if value is None:
            return ""
        return str(value) 
    #bold = workbook.add_format({'bold': True})
    
    ws = wb[sheet_names[0]]
    #make default
    for row in ws.rows:
        for cell in row:
            cell.font = Font(name='Calibri',
                        size=11,
                        bold=False,
                        italic=False,
                        vertAlign=None,
                        underline='none',
                        strike=False,
                        color='FF000000')
            cell.fill = PatternFill(fill_type=None,
                        start_color='FFFFFFFF',
                        end_color='FF000000')
            cell.border = Border(left=Side(border_style=None,
                                           color='FF000000'),
                                right=Side(border_style=None,
                                           color='FF000000'),
                                top=Side(border_style=None,
                                         color='FF000000'),
                                bottom=Side(border_style=None,
                                            color='FF000000'),
                                diagonal=Side(border_style=None,
                                              color='FF000000'),
                                diagonal_direction=0,
                                outline=Side(border_style=None,
                                             color='FF000000'),
                                vertical=Side(border_style=None,
                                              color='FF000000'),
                                horizontal=Side(border_style=None,
                                               color='FF000000')
                                                           )
            cell.alignment=Alignment(horizontal='general',
                        vertical='bottom',
                        text_rotation=0,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0)
            
            
            #cell.number_format = 'General'
            #cell.protection = Protection(locked=True,
                           # hidden=False)
            
            
    nciColor = 'FF'+ '%02x%02x%02x' % (131, 36, 52)
    for cell in ws.rows[0]:
        cell.font = Font(bold=True,  color='FFFFFFFF')
        cell.fill = PatternFill(fill_type='solid',
                        start_color=nciColor,
                        end_color=nciColor)
        
  
    row_count = ws.max_row
    column_count = ws.max_column

    #row = ws.row_dimensions[1]
    #row.font = Font(italic=True)
    for column_cells in ws.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column].width = length
    
    dims = {}
    colWidth = {}
    for row in ws.rows:
        for cell in row:
            colWidth[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value
    
    #col = ws.column_dimensions['A']
    #col.font = Font(bold=True)
    
    wb.save(filename = '%s/%s.xlsx' % (DIR, outFile))
    
   

openpyxlFormat('H:/', 'WV actions', 'WV actions2')              


def confirm(prompt=None, resp=False):

    
    if prompt is None:
        prompt = 'Confirm'

    if resp:
        prompt = '%s %s|%s: ' % (prompt, 'y', 'n')
    else:
        prompt = '%s %s|%s: ' % (prompt, 'n', 'y')
        
    while True:
        ans = input(prompt)
        if not ans:
            return resp
        if ans not in ['y', 'Y', 'n', 'N']:
            print( 'please enter y or n.')
            continue
        if ans == 'n' or ans == 'N':
            return 'True'
        if ans == 'Y' or ans == 'y':
            return input('please type the names of the %s variables, you want, \
            separated by a comma.' % len('jyason')) 
            
answer = confirm('Hello, %s are the current column names. Would you like to \
    change them?' % 'jahyson')           


if answer == 'True':
    columns = id_names
else:
    columns = answer.split(',')

   
    



def writerFormat(DIR, inFile, outFile='Same'):
    if outFile == 'Same':
        outFile = inFile
    # import pandas
    import pandas as pd
    
    # create writer
    
    # outFile format check
    
    if outFile[-5:] == '.xlsx':
        outFile = outFile[:-5] 
    elif outFile[-4:] == '.xls':
        outFile = outFile[:-4] 
    elif '.' in outFile:
        outFile = outFile[:outFile.find('.')]
        print("User Warning: Program removed 'outFile' type and replaced with '.xlsx'. \
              Sorry for your inconvenience"  )   
        
    writer = pd.ExcelWriter('%s/%s.xlsx' % (DIR, outFile), engine='xlsxwriter')

    # create xls object that contains all sheets of Excel file
    if '.' not in inFile:
        try:
            xls = pd.ExcelFile('%s/%s.xlsx' % (DIR, inFile))
            print("The 'inFile' worked using an xlsx format")
        except FileNotFoundError:
            try:
                xls = pd.ExcelFile('%s/%s.xls' % (DIR, inFile))
                print("The 'inFile' worked using an xls format")

            except FileNotFoundError:
                print("The 'inFile' type seems not to be compatible or non-existent.         \
                      Please make sure the 'inFile' file exist and/ or make sure             \
                      it is an xls or xlsx file and try again")
            else:
                xls = pd.ExcelFile('%s/%s.xls' % (DIR, inFile))
                sufix = '.xls' 
        else:
            xls = pd.ExcelFile('%s/%s.xlsx' % (DIR, inFile)) 
            sufix = '.xlsx' 


        
    elif '.' in inFile:
        tempSufix = inFile[-4:]
        inFile = inFile[:inFile.find('.')]
        try:
            xls = pd.ExcelFile('%s/%s.xlsx' % (DIR, inFile))
            if tempSufix != 'xlsx':
                print("The 'inFile' worked, but only after using an xlsx format")
        except FileNotFoundError:
            try:
                xls = pd.ExcelFile('%s/%s.xls' % (DIR, inFile))
                if tempSufix != '.xls':
                    print("The 'inFile' worked, but only after using an xls format")
            except FileNotFoundError:
                print("The 'inFile' type seems not to be compatible or non-existent.         \
                      Please make sure the 'inFile' file exist and/ or make sure             \
                      it is an xls or xlsx file and try again")
            else:
                xls = pd.ExcelFile('%s/%s.xls' % (DIR, inFile))
                sufix = '.xls' 
        else:
            xls = pd.ExcelFile('%s/%s.xlsx' % (DIR, inFile))
            sufix = '.xlsx' 

     
    # get sheet names of workbook to use
    worksheets = xls.sheet_names
    
    print(worksheets)
    
    # iterate through all sheets
    for i, sheet in enumerate(worksheets):
      
        # create dataframe from excel file
        df = pd.read_excel('%s/%s%s' % (DIR, inFile, sufix), sheetname=sheet)
        
        valLen = []
        for i in df:
            valLen.append(int(df[i].map(lambda x: len(str(x))).max()))
        
        origCol = []
        newCol = []
        colCount = 0
        for col in df:
            colCount += 1
            origCol.append(str(col)) 
            newCol.append('col'+str(colCount))
        
        colLen = []
        for i in range(len(valLen)):
            colLen.append(max(valLen[i],len(origCol[i])))
           
        # write file to outDIR renamed
        df.to_excel(writer, sheet_name=sheet, startrow = 1, header = False, index = False)
        
        workbook = writer.book
           
        worksheet = writer.sheets[sheet]
        
        worksheet.autofilter(0, 0, df.shape[0], df.shape[1])
        
        
        for i in range(len(colLen)):
            worksheet.set_column(i, i, colLen[i])
            
        hex = '#%02x%02x%02x' % (131, 36, 52)
        
        header_format = workbook.add_format({
            'bold': True,
            'font_color' : '#ffffff',
            'text_wrap': False,
            'valign': 'center',
            'align' : 'center',
            'fg_color': hex,
            'border': 0})
        
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
         
            
    writer.close()    
        
writerFormat('H:/', 'test excel2')





def excelAutoWidth(DIR, file):
    import pandas as pd
    
    # create dataframe from excel file
    df = pd.read_excel('%s/%s' % (DIR, file))
    
    valLen = []
    for i in df:
        valLen.append(int(df[i].map(lambda x: len(str(x))).max()))
    
    origCol = []
    newCol = []
    colCount = 0
    for col in df:
        colCount += 1
        origCol.append(str(col)) 
        newCol.append('col'+str(colCount))
    
    colLen = []
    for i in range(len(valLen)):
        colLen.append(max(valLen[i],len(origCol[i])))
       
        
    
    # create writer
    writer = pd.ExcelWriter('%s/%s.xlsx' % (DIR, file), engine='xlsxwriter')
    # write file to outDIR renamed

   # write file to outDIR renamed
    df.to_excel(writer, index = False)



    
def nciHeaders(DIR, file):
    
    df = pd.read_excel('%s/%s' % (DIR, file))

    
    writer = pd.ExcelWriter('%s/%s.xlsx' % (DIR, outFile), engine='xlsxwriter')
    # write file to outDIR renamed
    df.to_excel(writer,"Sheet1", startrow = 1, header = False, index = False)
    
    workbook = writer.book
    
    sheetList = []
    for worksheet in workbook.worksheets():
        sheetList.append(worksheet.get_name())
    
        
    worksheet = writer.sheets[sheetList[0]]
    
    worksheet.autofilter(0, 0, df.shape[0], df.shape[1])
    
    
    for i in range(len(colLen)):
        worksheet.set_column(i, i, colLen[i])
        
    hex = '#%02x%02x%02x' % (131, 36, 52)
    
    header_format = workbook.add_format({
        'bold': True,
        'font_color' : '#ffffff',
        'text_wrap': False,
        'valign': 'center',
        'align' : 'center',
        'fg_color': hex,
        'border': 0})
    
    for col_num, value in enumerate(df.columns.values):
        worksheet.write(0, col_num, value, header_format)
     
        
    writer.close()    

DIR = 'H:/'
inFile = 'test excel.xlsx'
outFile =  'test excel2' 

import pandas as pd


def autoFilter(DIR, inFile, outFile):
    df = []

    xls = pd.ExcelFile('%s/%s' % (DIR, inFile))
    worksheets = xls.sheet_names
    writer = pd.ExcelWriter('%s/%s.xlsx' % (DIR, outFile), engine='xlsxwriter')
    
    for i, sheet in enumerate(worksheets):
        df = (pd.read_excel('%s/%s' % (DIR, inFile), sheetname=sheet))
        df.to_excel(writer, sheet_name=sheet, index = False)
    writer.close()

    workbook = writer.book
    
    sheetList = []
    for worksheet in workbook.worksheets():
        sheetList.append(worksheet.get_name())
    
    for i, sheet in enumerate(sheetList):
        sheet.autofilter(0, 0, df.shape[0], df.shape[1])

    

    
    print(df.shape[0])
        
    worksheet = writer.sheets[sheetList[0]]
    
    worksheet.autofilter(0, 0, df.shape[0], df.shape[1])
    
    sheetList = []
    for worksheet in workbook.worksheets():
        sheetList.append(worksheet.get_name())
    
    for sheet in enumerate(sheetList):    
        
        
    
    
autoFilter('H:/', 'test excel.xlsx', 'test excel2')
    
workbook = writer.book  
    

soup2= str(soup)
chars = """`'"*_{}[]()<>&#+.!$%@=:;,"""
for c in chars:
    soup2 = soup2.replace(c, " ")
listCodes = []  
for word in text.split():
    print(word)

print(soup)

for i in soup:
    print(i.text)
    print(i.attrs)
    print(i.tagstack)
    
    
    
    
    
    
    
    
    
    
    
# -*- coding: utf-8 -*-
"""
Created on Tue Aug 15 14:37:00 2017

@author: backesj
"""

import requests
from bs4 import BeautifulSoup as bs4
import pandas as pd

#set parameters

url = 'http://public.oed.com/how-to-use-the-oed/abbreviations/'
outDIR = 'S:\\DA_work_files\\DA_Work_Jayson\\NC_Provider_Scrape.xlsx'


#run requests to get url content

page_content = requests.get(url).content

#create soup object using beautifulsoup

soup = bs4(page_content, "html.parser")





header_list = []
text_list = []
counter = 0 
list1 = []



abr_dict = {}
for table in soup.find_all('tbody'):
    for td in table.find_all('td'):
        print(td.text)
        print(td)

        if td.attrs != {'colspan': '2'}:
            temp = td.text.strip().lower()
            temp = ''.join(char for char in temp if char not in set("/.,()'"))
            if len(temp) > 0:
                list1.append(temp)
abr_dict = dict(list1[i:i+2] for i in range(0, len(list1), 2))
abr_dict2 = {y:x for x,y in abr_dict.items()}
             
             
abr_dict3 = {k: v for k, v in abr_dict2.items() if 'in' not in v}
abr_dict4 = {k: v for k, v in abr_dict3.items() if ' ' not in k}

academia academy academical
test_list = ['academia','academy','jayson']  
  
for i in test_list:
    for key, value in abr_dict.items():
        if i in value:
            print(key)


            
            
            
            
for word in test_list:
    if word in abr_dict[value]:
            print(abr_dict[key])
if test_str != 'test':
    print('jayson')
        
        
        

            if len(td.text) > 1: 
            print(td.text)
    for a in td.find_all('a'):
        if a.text != 'PDF':
            if 'header' in a['href']:
                header_list.append(url[0:29] + a['href'])
            else:
                text_list.append(url[0:29] +a['href'])
                
                
                
def dedup(seq):
    """
    removes duplicate values from a string or list
    """
    seen = set()
    seen_add = seen.add
    return [x for x in seq if not (x in seen or seen_add(x))]    

            
            
            
# -*- coding: utf-8 -*-
"""
Created on Mon Sep 18 09:41:43 2017

@author: backesj
"""



def function1():
    print ("called function 1")

def function2():
    print( "called function 2")

def function3():
    print ("called function 3")

tokenDict = {"cat":function1, "dog":function2, "bear":function3}

# simulate, say, lines read from a file
lines = ["cat","bear","cat","dog"]

for line in lines:
    
    # lookup the function to call for each line
    functionToCall = tokenDict[line]

    # and call it
    functionToCall()


import pandas as pd

DIR = 'H:/Python Webscraping/NC/Results'

inFile = 'NC_Provider_Scrape'

xls = pd.ExcelFile('%s/%s.xlsx' % (DIR, inFile))

worksheets = xls.sheet_names

print(worksheets)

column_list = []
# iterate through all sheets
for i, sheet in enumerate(worksheets):
  
    # create dataframe from excel file
    df = pd.read_excel('%s/%s.xlsx' % (DIR, inFile), sheetname=sheet)
    for column in df:
        column_list.append(column)
 
        
        
        
def dedup(seq):
    """
    removes duplicate values from a list or 
    duplicate characters from a string 
    """
    if type(seq) == list:
        seen = set()
        seen_add = seen.add
        return [x for x in seq if not (x in seen or seen_add(x))]    
    elif type(seq) == str:
        seen = set()
        seen_add = seen.add
        return ''.join([x for x in seq if not (x in seen or seen_add(x))])    
    else:
        print("Currently function can only handle lists and strings")
        
column_list = dedup(column_list)

import pandas as pd
df = pd.read_excel('H:/Python/Dictionaries/State List.xlsx', header=None, index_col=0 ) 

import numpy

state_dict = {}

for j in range((len(df.index))):
    for i in range(3):
        state_dict[df[i+1][j]] = df.index[j]





print(state_dict['ala'])

















key_list = []



for i in df:
    print(df[i]) + df[1:5]
    for j in range((len(df.index))):
        print(i[0] + j)

for i in range((len(df.index))):
    for j in df[i]:
        
        print(df[i] + j)
    
    
    print(i)
for i in df:
    for j in i:
       print(j)




df2 = df.pivot(index = 0, columns = 1,2,3, values='value')
df = pd.read_excel('H:/Python/Dictionaries/State List.xlsx', header=None) 
state_dict = {}
for i in df.iterrows():
    print(i)
    state_dict[i[1][0]] = list(i[1][1:5])

jay = ['ar', 'ala', 'dn', 'nd', 'alaska']


for i in jay:
    print([k for k, v in state_dict.items() if i == v])
            print(k)
        else:
            print('i ' + 'is not found')


for i in state_dict():
    print(i)
    print(i[key])
    fileType = ['H:/','S:/DA_work_files/DA_work_Jayson','C:/Users/backesj']

def getLatestFile(DIR, FileType):

    #import os packages
    import os
    
    from os import listdir
    from os.path import isfile, join
    file_list2 = []
    folder_list2 = []
    for i in fileType:
    #import platform
        folder_list = [item for item in os.listdir(i) if os.path.isdir(os.path.join(i, item))]
        print(folder_list)
        file_list = [f for f in listdir(i) if isfile(join(i, f)) if '.py' in f]
        file_list2.append(file_list)
        folder_list2.append(folder_list)
    # create file list of xls files in download file directory
            for file in file_list2:
                # get date of file creation in download file directory
                file_list2.append(os.path.getmtime('%s/%s' % (downloadDIR, file)))
    # create empty dictionary
    d = {}
    # create dictionary of files and date 
    for i in range(len(file_list)):
        d[file_list[i]] = file_list2[i]

    # get max item in dictionary (last item created) this should be the file you
    # downloaded from the website
    import operator
    return max(d.items(), key=operator.itemgetter(1))[0]

with open('H:\Python Webscraping\python soup cheatsheet v1.1.py', 'r') as file:
    data = file.read()
file.close()

import glob
import os

for i in fileType:
    print(glob.glob('%s/*.py' % i))

    
    
 
    
def findText(inList, fileType, searchTerm):
    """
    This function takes a directory or a list of directories and searches all folders
    for a type of file using set of text that the user defines
    """
    if type(inList) == str:
        newList = inList.split()
    else:
        newList = inList
    import os
        
    from os import listdir
    from os.path import isfile, join
    outList = []
    for i in newList:
        print(i)
        for root, dirs, files in os.walk(i):
            for file in files:
                if file.endswith(fileType):
                    if fileType == '.sas':
                        with open(os.path.join(root, file),'r', encoding = 'latin1') as text:
                            data = text.read()
                            if searchTerm in data:
                                outList.append(os.path.join(root, file))
                                text.close()
                            else:
                                text.close()
                    else:
                        with open(os.path.join(root, file),'r') as text:
                            data = text.read()
                            if searchTerm in data:
                                outList.append(os.path.join(root, file))
                                text.close()
                            else:
                                text.close()
    return outList
fileList = findText( ["C:/Users/backesj", "H:/", "S:/DA_work_files/DA_work_jayson"],
                    ".py", 
                    "sas7bdat") 
                    
                    
                    
                    
                 ["C:/Users/backesj", "H:/", "S:/DA_work_files/DA_work_jayson"]
            
# yes no workings




def confirm(prompt=None, resp=False):
    """prompts for yes or no response from the user. Returns True for yes and
    False for no.

    'resp' should be set to the default value assumed by the caller when
    user simply types ENTER.

    >>> confirm(prompt='Create Directory?', resp=True)
    Create Directory? [y]|n: 
    True
    >>> confirm(prompt='Create Directory?', resp=False)
    Create Directory? [n]|y: 
    False
    >>> confirm(prompt='Create Directory?', resp=False)
    Create Directory? [n]|y: y
    True

    """
    
    if prompt is None:
        prompt = 'Confirm'

    if resp:
        prompt = '%s [%s]|%s: ' % (prompt, 'y', 'n')
    else:
        prompt = '%s [%s]|%s: ' % (prompt, 'n', 'y')
        
    while True:
        ans = input(prompt)
        if not ans:
            return resp
        if ans not in ['y', 'Y', 'n', 'N']:
            print('please enter y or n.')
            continue
        if ans == 'y' or ans == 'Y':
            return True
        if ans == 'n' or ans == 'N':
            return False
            
            
            
# multiple choice

     

def choose(prompt=None, resp=False):
    """prompts for yes or no response from the user. Returns True for yes and
    False for no.

    'resp' should be set to the default value assumed by the caller when
    user simply types ENTER.

    >>> confirm(prompt='Create Directory?', resp=True)
    Create Directory? [y]|n: 
    True
    >>> confirm(prompt='Create Directory?', resp=False)
    Create Directory? [n]|y: 
    False
    >>> confirm(prompt='Create Directory?', resp=False)
    Create Directory? [n]|y: y
    True

    """

    if prompt is None:
        prompt = 'Confirm'

    if resp:
        prompt = '%s %s: ' % (prompt, choices)
    else:
        prompt = '%s %s: ' % (prompt, choices)
        
    while True:
        ans = input(prompt)
        if not ans:
            return resp
        if ans == '0':
            return False
        try:
            int(ans)
        except ValueError:
            print('***** ERROR: please enter an integer of %s, Or enter 0 to exit *****' % choices)
        else:
            if int(ans) not in numList:
                print('***** ERROR: please enter %s. Or enter 0 to exit *****' % choices)
                continue
            if int(ans) in numList:
                while True:
                    ans2 = input('you chose %s, are you sure [y/n]' % j[ans])
                    if ans2 not in ['y', 'Y', 'n', 'N']:
                        print('please enter y or n.')
                        continue
                    if ans2 == 'y' or ans2 == 'Y':
                        return drAddresses[str(int(ans)-1)]['link']
                    if ans2 == 'n' or ans2 == 'N':
                        break

# number of params, works with choose multiple choice

def numParams(inDict):
    
    numList = []
    firstNums = []
    for i in range(len(inDict)):
        numList.append(i+1)
        if i+1 != len(inDict):
            firstNums.append(str(i+1))
        else:
            lastNum = str(i+1)
    if len(inDict) > 2:
        choices = ', '.join(firstNums) + ', or ' + lastNum
    else:
        choices = ', '.join(firstNums) + ' or ' + lastNum  
    return numList, choices
    
               
numList, choices = numParams(matches)
    
import pandas as pd

answerDF = pd.DataFrame.from_dict(matches, orient='index')
answerDF.columns = ['Address', 'Score']
 

# os start 

os.system(""" "start chrome.exe "http://dph1.adph.state.al.us/FacilitiesDirectory" """)





OPTION EMAILAUTHPROTOCOL=NONE EMAILHOST="BALT-EMAIL1.admedcorp.com" EMAILID="BACKESJ@ADMEDCORP.COM" Emailsys=SMTP ;
FILENAME Mailbox EMAIL;
DATA _NULL_;
FILE Mailbox TO='6158098494@vtext.com';
PUT "REDO DONE";
RUN;


import smtplib
smtpObj = smtplib.SMTP('BALT-EMAIL1.admedcorp.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('BACKESJ@ADMEDCORP.COM', '')
smtpObj.sendmail('BACKESJ@ADMEDCORP.COM', 'BACKESJ@ADMEDCORP.COM', 'Subject: So long')
                 Dear Alice, so long and thanks for all the fish. Sincerely, Bob')

    
tempText = 'jayson an backes the AR ' 
if any(substring in tempText for substring in zone):
    print('yes')   